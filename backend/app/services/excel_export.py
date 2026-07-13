"""
Excel Export Service — Phase 6

Uses openpyxl to append confirmed student grades and marks details
into a single consolidated 'exports/marks.xlsx' spreadsheet.
"""
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from app.models.schemas import SubmissionRecord

EXPORT_DIR = Path(__file__).parent.parent.parent / "exports"
EXPORT_FILE = EXPORT_DIR / "marks.xlsx"


def append_submission_to_excel(record: SubmissionRecord) -> str:
    """
    Appends one row representing the confirmed student grades to the marks.xlsx spreadsheet.
    Creates the workbook with correct column headers if it doesn't already exist.
    """
    if not record.extraction_result:
        raise ValueError("Cannot export submission: extraction_result is empty.")

    # Create exports directory if it doesn't exist
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    headers = [
        "Name",
        "Roll No",
        "Branch",
        "Subject",
        "Date",
        "Marks Breakdown",
        "Total Marks",
        "Validation Status",
    ]

    # Load or create workbook
    if not EXPORT_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Marks Records"
        ws.append(headers)
    else:
        wb = load_workbook(EXPORT_FILE)
        ws = wb.active

    # Format the Marks Breakdown string (e.g. "1a-5, 1b-5, 2a-4")
    ext = record.extraction_result
    marks_str = ", ".join(
        f"{entry.question_no}{entry.part}-{entry.marks}"
        for entry in sorted(ext.marks_entries, key=lambda x: (x.question_no, x.part))
    )

    # Determine validation status
    val_status = "Discrepancy noted"
    if record.validation_result and record.validation_result.overall_status == "valid":
        val_status = "Valid"

    # Assemble row data
    row_data = [
        ext.name,
        ext.roll_no,
        ext.branch,
        ext.subject,
        ext.date,
        marks_str,
        ext.total_marks_declared,
        val_status,
    ]

    # Append and save
    ws.append(row_data)
    wb.save(EXPORT_FILE)
    return str(EXPORT_FILE)
